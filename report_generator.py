import webbrowser
import hashlib
import base64
import os
import time
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

from urllib.parse import urlencode, parse_qs
from http.server import HTTPServer, BaseHTTPRequestHandler
import threading
import ssl
from datetime import datetime, timedelta
import json
from collections import defaultdict
import gspread
from gspread_formatting import *
import openpyxl
from openpyxl.styles import PatternFill, Font

# -- ОБЩИЕ НАСТРОЙКИ --
CLIENT_ID = "c7z6p"
CLIENT_SECRET = "eDWm62flZAqlvk4GGkf7xMXlfyCIPPil"
REDIRECT_URI = "https://localhost:5001"
SCOPES = "openid offline_access shared incentives"

# -- НАСТРОЙКИ GOOGLE SHEETS --
GOOGLE_CREDENTIALS_FILE = 'google_credentials.json'
GOOGLE_SHEET_NAME = 'Dodo Reports'  # <-- ЗАМЕНИТЕ НА НАЗВАНИЕ ВАШЕЙ ТАБЛИЦЫ

# -- ИМЕНА ФАЙЛОВ --
DODO_TOKEN_FILE = 'dodo_token_reports.json'
PIZZERIAS_FILE = 'ID пиццерий.txt'
SALES_DATA_FILE = 'sales_data.json'
REWARDS_DATA_FILE = 'rewards_data.json'
OUTPUT_REPORT_FILE = 'labour_cost_report.xlsx'
VERIFICATION_REPORT_FILE = 'Проверка выручки и зарплаты для LC.txt'
REQUEST_FILE = 'report_request.json'
STAFF_ROLE_CACHE_FILE = 'staff_role_cache.json'

# -- Глобальные переменные --
auth_event = threading.Event()
authorization_code = None
dodo_tokens = {}


# --- БЛОК 1: АВТОРИЗАЦИЯ DODO IS API ---

def save_data(filename, data):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


def save_dodo_tokens(tokens):
    global dodo_tokens
    dodo_tokens = tokens
    save_data(DODO_TOKEN_FILE, tokens)


def load_dodo_tokens():
    global dodo_tokens
    if os.path.exists(DODO_TOKEN_FILE):
        with open(DODO_TOKEN_FILE, 'r') as f: dodo_tokens = json.load(f)
        return True
    return False


def refresh_dodo_token():
    if 'refresh_token' not in dodo_tokens: return False
    print("[Генератор] [Токен] Попытка обновить токен...")
    data = {'grant_type': 'refresh_token', 'refresh_token': dodo_tokens['refresh_token'], 'client_id': CLIENT_ID,
            'client_secret': CLIENT_SECRET}
    try:
        response = requests.post("https://auth.dodois.io/connect/token", data=data,
                                 headers={'Content-Type': 'application/x-www-form-urlencoded'}, verify=False)
        response.raise_for_status()
        save_dodo_tokens(response.json())
        print("[Генератор] [Токен] Токен успешно обновлен.")
        return True
    except requests.RequestException:
        if os.path.exists(DODO_TOKEN_FILE): os.remove(DODO_TOKEN_FILE)
        return False


def get_access_token():
    if not load_dodo_tokens() or not refresh_dodo_token():
        if not perform_full_auth_flow(): return None
    return dodo_tokens.get('access_token')


def perform_full_auth_flow():
    global authorization_code
    authorization_code = None
    auth_event.clear()

    try:
        server = HTTPServer(('localhost', 5001), CallbackHandler)
        context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        context.load_cert_chain(certfile='cert.pem', keyfile='key.pem')
        server.socket = context.wrap_socket(server.socket, server_side=True)
    except FileNotFoundError:
        print("\n[Критическая ошибка] Не найдены файлы сертификата 'cert.pem' и 'key.pem'.")
        return False

    server_thread = threading.Thread(target=server.serve_forever, daemon=True)
    server_thread.start()

    code_verifier, code_challenge = generate_pkce_codes()
    auth_params = {'client_id': CLIENT_ID, 'scope': SCOPES, 'response_type': 'code', 'redirect_uri': REDIRECT_URI,
                   'code_challenge': code_challenge, 'code_challenge_method': 'S256'}
    auth_url = f"https://auth.dodois.io/connect/authorize?{urlencode(auth_params)}"
    print(f"\n[АВТОРИЗАЦИЯ] Требуется ручной вход. Пожалуйста, откройте эту ссылку в браузере:\n{auth_url}\n")

    auth_event.wait(timeout=300)
    server.shutdown()

    if not authorization_code:
        print("[Авторизация] Ошибка: не удалось получить код авторизации.")
        return False

    try:
        token_data = {'client_id': CLIENT_ID, 'client_secret': CLIENT_SECRET, 'grant_type': 'authorization_code',
                      'code': authorization_code, 'code_verifier': code_verifier, 'redirect_uri': REDIRECT_URI,
                      'scope': SCOPES}
        response = requests.post("https://auth.dodois.io/connect/token", data=token_data,
                                 headers={'Content-Type': 'application/x-www-form-urlencoded'}, verify=False)
        response.raise_for_status()
        save_dodo_tokens(response.json())
        print("[Авторизация] Успешно! Токены сохранены.")
        return True
    except requests.exceptions.HTTPError as e:
        print(f"\n[Критическая ошибка авторизации] {e.response.text}")
        return False


def generate_pkce_codes():
    code_verifier = base64.urlsafe_b64encode(os.urandom(40)).decode('utf-8').rstrip('=')
    challenge_bytes = hashlib.sha256(code_verifier.encode('utf-8')).digest()
    code_challenge = base64.urlsafe_b64encode(challenge_bytes).decode('utf-8').rstrip('=')
    return code_verifier, code_challenge


class CallbackHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        global authorization_code
        if '?' in self.path and 'code' in self.path:
            parsed_path = parse_qs(self.path.split('?', 1)[1])
            authorization_code = parsed_path['code'][0]
            self.send_response(200)
            self.send_header("Content-type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write("<h1>Авторизация прошла успешно!</h1><p>Можете закрыть это окно.</p>".encode('utf-8'))
            auth_event.set()


# --- БЛОК 2: ФУНКЦИИ СБОРА ДАННЫХ ---
def fetch_sales_data(access_token, unit_ids, from_date, to_date):
    print("  [API] Запрашиваем данные о продажах...")
    all_sales_data = []
    for i in range(0, len(unit_ids), 30):
        chunk = unit_ids[i:i + 30]
        params = {'units': ",".join(chunk), 'from': from_date, 'to': to_date}
        response = requests.get("https://api.dodois.io/dodopizza/ru/finances/sales/units",
                                headers={'Authorization': f'Bearer {access_token}'}, params=params, verify=False)
        if response.status_code == 200:
            all_sales_data.extend(response.json().get('result', []))
        else:
            raise Exception(f"Ошибка API продаж: {response.text}")
    save_data(SALES_DATA_FILE, {"result": all_sales_data})
    return {"result": all_sales_data}


def fetch_rewards_data(access_token, unit_ids, from_date, to_date, is_for_cache=False):
    print(f"  [API] Запрашиваем данные о вознаграждениях ({'для кэша' if is_for_cache else 'для отчета'})...")
    params = {'units': ",".join(unit_ids), 'from': from_date, 'to': to_date}
    response = requests.get("https://api.dodois.io/dodopizza/ru/staff/incentives-by-members",
                            headers={'Authorization': f'Bearer {access_token}'}, params=params, verify=False)
    if response.status_code == 200:
        rewards_data = response.json()
        if not is_for_cache:
            save_data(REWARDS_DATA_FILE, rewards_data)
        return rewards_data
    else:
        raise Exception(f"Ошибка API вознаграждений: {response.text}")


# --- БЛОК 3: ФУНКЦИИ ФОРМИРОВАНИЯ ОТЧЕТА ---
def load_staff_role_cache():
    if os.path.exists(STAFF_ROLE_CACHE_FILE):
        with open(STAFF_ROLE_CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def update_and_save_staff_role_cache(rewards_data_for_cache):
    print("  [Кэш] Обновляем кэш должностей сотрудников...")
    role_cache = load_staff_role_cache()
    for member in rewards_data_for_cache.get('staffMembers', []):
        staff_id = member.get('staffId')
        shifts = member.get('shiftsDetailing', [])
        if staff_id and shifts:
            known_types = [s.get('staffType') for s in shifts if
                           s.get('staffType') and s.get('staffType') != 'Неизвестный тип']
            if known_types:
                primary_role = max(set(known_types), key=known_types.count)
                role_cache[staff_id] = primary_role

    save_data(STAFF_ROLE_CACHE_FILE, role_cache)
    print(f"  [Кэш] Кэш должностей успешно обновлен. Всего в кэше {len(role_cache)} сотрудников.")


def generate_and_save_report(pizzerias, sales_data, rewards_data):
    print("  [Отчет] Формируем сводный отчет...")
    role_cache = load_staff_role_cache()

    report_data = defaultdict(lambda: {
        'total_revenue': 0, 'delivery_revenue': 0, 'kitchen_reward': 0, 'courier_reward': 0
    })

    for p_id, p_name in pizzerias.items():
        report_data[p_id]['name'] = p_name

        sales_for_unit = next((s for s in sales_data.get('result', []) if s['unitId'] == p_id), None)
        if sales_for_unit:
            report_data[p_id]['total_revenue'] = sales_for_unit.get('sales', 0)
            for breakdown in sales_for_unit.get('salesBreakdown', []):
                if breakdown.get('salesChannel') == 'Delivery':
                    report_data[p_id]['delivery_revenue'] += breakdown.get('sales', 0)

        for member in rewards_data.get('staffMembers', []):
            staff_id = member.get('staffId')
            shifts = member.get('shiftsDetailing', [])
            premiums = member.get('premiums', [])

            primary_role = role_cache.get(staff_id)
            if not primary_role and shifts:
                known_types = [s.get('staffType') for s in shifts if
                               s.get('staffType') and s.get('staffType') != 'Неизвестный тип']
                if known_types: primary_role = max(set(known_types), key=known_types.count)

            for shift in shifts:
                if shift.get('unitId') == p_id:
                    staff_type = shift.get('staffType')
                    if staff_type in ['KitchenMember', 'Cashier', 'PersonalManager']:
                        report_data[p_id]['kitchen_reward'] += shift.get('totalWage', 0)
                    elif staff_type == 'Courier':
                        report_data[p_id]['courier_reward'] += shift.get('totalWage', 0)

            for premium in premiums:
                if premium.get('unitId') == p_id:
                    if primary_role in ['KitchenMember', 'Cashier', 'PersonalManager']:
                        report_data[p_id]['kitchen_reward'] += premium.get('amount', 0)
                    elif primary_role == 'Courier':
                        report_data[p_id]['courier_reward'] += premium.get('amount', 0)

    # -- Промежуточные расчеты для сортировки и вывода --
    results_for_sorting = []
    for p_id, data in report_data.items():
        total_revenue_adjusted = data['total_revenue'] / 1.12
        delivery_revenue_adjusted = data['delivery_revenue'] / 1.12

        kitchen_cost_lc = data['kitchen_reward'] * 1.615731042
        courier_cost_lc = data['courier_reward']

        labour_cost_percent = ((
                                           kitchen_cost_lc + courier_cost_lc) / total_revenue_adjusted) * 100 if total_revenue_adjusted > 0 else 0
        labour_cost_percent += 2.52

        dc_percent = (data['courier_reward'] / delivery_revenue_adjusted) * 100 if delivery_revenue_adjusted > 0 else 0
        kc_percent = (kitchen_cost_lc / total_revenue_adjusted) * 100 if total_revenue_adjusted > 0 else 0

        if labour_cost_percent > 2.52:
            results_for_sorting.append({
                'name': data['name'], 'lc': labour_cost_percent,
                'dc': dc_percent, 'kc': kc_percent
            })

    sorted_results = sorted(results_for_sorting, key=lambda x: x['lc'])

    # -- Генерация отчетов --
    start_date = rewards_data.get('from', '')
    end_date = rewards_data.get('to', '')
    period_header = ""
    if start_date and end_date:
        start_date_fmt = datetime.fromisoformat(start_date).strftime('%d.%m.%Y')
        end_date_fmt = datetime.fromisoformat(end_date).strftime('%d.%m.%Y')
        period_header = f"{start_date_fmt} - {end_date_fmt}"

    generate_xlsx_report(sorted_results, period_header)
    send_report_to_google_sheet(sorted_results, period_header)
    generate_verification_report(pizzerias, report_data)


def generate_xlsx_report(sorted_results, period_header):
    print("  [Отчет] Создаем XLSX файл...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"LC Report {period_header}"

    header_font = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    header = ["Пиццерия", "LC %", "DC %", "KC %"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = header_font

    for item in sorted_results:
        row_data = [item['name'], item['lc'] / 100, item['dc'] / 100, item['kc'] / 100]
        ws.append(row_data)

        row_index = ws.max_row
        if item['lc'] < 27:
            fill = green_fill
        elif 27 <= item['lc'] < 30:
            fill = yellow_fill
        else:
            fill = red_fill
        for col in range(1, 5):
            ws.cell(row=row_index, column=col).fill = fill

    if sorted_results:
        avg_lc = sum(item['lc'] for item in sorted_results) / len(sorted_results)
        avg_dc = sum(item['dc'] for item in sorted_results) / len(sorted_results)
        avg_kc = sum(item['kc'] for item in sorted_results) / len(sorted_results)
        ws.append([])
        avg_row = ["Среднее значение по сети", avg_lc / 100, avg_dc / 100, avg_kc / 100]
        ws.append(avg_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    for col_letter in ['B', 'C', 'D']:
        for cell in ws[col_letter]:
            cell.number_format = '0.00"%"'

    ws.column_dimensions['A'].width = 35
    for col_letter in ['B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 12

    wb.save(OUTPUT_REPORT_FILE)
    print(f"  [Отчет] XLSX отчет сохранен в '{OUTPUT_REPORT_FILE}'.")


def send_report_to_google_sheet(sorted_results, period_header):
    try:
        print("  [Google Sheets] Попытка отправки отчета...")
        gc = gspread.service_account(filename=GOOGLE_CREDENTIALS_FILE)
        sh = gc.open(GOOGLE_SHEET_NAME)

        worksheet_title = "Labour Cost Report"
        try:
            worksheet = sh.worksheet(worksheet_title)
        except gspread.WorksheetNotFound:
            worksheet = sh.add_worksheet(title=worksheet_title, rows="100", cols="20")

        worksheet.clear()

        header = ["Пиццерия", "LC %", "DC %", "KC %"]
        rows_to_send = [header]
        for item in sorted_results:
            rows_to_send.append([item['name'], item['lc'] / 100, item['dc'] / 100, item['kc'] / 100])

        if sorted_results:
            avg_lc = sum(item['lc'] for item in sorted_results) / len(sorted_results)
            avg_dc = sum(item['dc'] for item in sorted_results) / len(sorted_results)
            avg_kc = sum(item['kc'] for item in sorted_results) / len(sorted_results)
            rows_to_send.append([])
            rows_to_send.append(["Среднее значение по сети", avg_lc / 100, avg_dc / 100, avg_kc / 100])

        worksheet.update('A1', rows_to_send, value_input_option='USER_ENTERED')

        worksheet.format("B:D", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
        set_frozen(worksheet, rows=1)

        green = Color(0.776, 0.937, 0.804)
        yellow = Color(1.0, 0.922, 0.612)
        red = Color(1.0, 0.78, 0.8)

        rule1 = ConditionalFormatRule(
            ranges=[GridRange.from_a1_range(f'A2:D{len(rows_to_send)}', worksheet)],
            booleanRule=BooleanCondition('NUMBER_LESS_THAN', [0.27], relative=False),
            format=CellFormat(backgroundColor=green)
        )
        rule2 = ConditionalFormatRule(
            ranges=[GridRange.from_a1_range(f'A2:D{len(rows_to_send)}', worksheet)],
            booleanRule=BooleanCondition('NUMBER_BETWEEN', [0.27, 0.30], relative=False),
            format=CellFormat(backgroundColor=yellow)
        )
        rule3 = ConditionalFormatRule(
            ranges=[GridRange.from_a1_range(f'A2:D{len(rows_to_send)}', worksheet)],
            booleanRule=BooleanCondition('NUMBER_GREATER_THAN_OR_EQUAL_TO', [0.30], relative=False),
            format=CellFormat(backgroundColor=red)
        )

        rules = get_conditional_format_rules(worksheet)
        rules.clear()
        rules.append(rule1)
        rules.append(rule2)
        rules.append(rule3)
        rules.save()

        worksheet.update_title(f"LC Report ({period_header})")
        print("  [Google Sheets] Отчет успешно отправлен и отформатирован.")

    except gspread.exceptions.SpreadsheetNotFound:
        print(f"  [Google Sheets][Ошибка] Таблица с именем '{GOOGLE_SHEET_NAME}' не найдена.")
    except FileNotFoundError:
        print(f"  [Google Sheets][Ошибка] Файл ключей '{GOOGLE_CREDENTIALS_FILE}' не найден.")
    except Exception as e:
        print(f"  [Google Sheets][Ошибка] Произошла непредвиденная ошибка: {e}")


def generate_verification_report(pizzerias, report_data):
    verification_lines = []
    header_ver = " Проверка выручки и зарплаты для LC "
    verification_lines.append(header_ver.center(80, '=') + "\n\n")
    for p_id, p_name in sorted(pizzerias.items(), key=lambda item: item[1]):
        data = report_data[p_id]
        total_revenue_adjusted = data['total_revenue'] / 1.12
        delivery_revenue_adjusted = data['delivery_revenue'] / 1.12
        verification_lines.append(f"🍕 Пиццерия: {p_name}\n")
        verification_lines.append(f"   - Общая выручка: {data['total_revenue']:,.2f} руб.\n")
        verification_lines.append(f"   - Общая выручка (для расчета): {total_revenue_adjusted:,.2f} руб.\n")
        verification_lines.append(f"   - Выручка доставки: {data['delivery_revenue']:,.2f} руб.\n")
        verification_lines.append(f"   - Выручка доставки (для расчета): {delivery_revenue_adjusted:,.2f} руб.\n")
        verification_lines.append(f"   - Зарплата кухни (вкл. кассиров): {data['kitchen_reward']:,.2f} руб.\n")
        verification_lines.append(f"   - Зарплата курьеров: {data['courier_reward']:,.2f} руб.\n")
        verification_lines.append("-" * 40 + "\n")

    with open(VERIFICATION_REPORT_FILE, 'w', encoding='utf-8') as f:
        f.writelines(verification_lines)
    print(f"  [Отчет] Отчет для проверки сохранен в файл '{VERIFICATION_REPORT_FILE}'.")


def load_pizzerias():
    if not os.path.exists(PIZZERIAS_FILE): return None
    pizzerias_map = {}
    with open(PIZZERIAS_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            if ' - ' in line:
                p_id, p_name = line.strip().split(' - ', 1)
                pizzerias_map[p_id] = p_name
    return pizzerias_map


# --- ГЛАВНАЯ ФУНКЦИЯ (ДЕМОН) ---
def main():
    print("[Генератор] Запущен в режиме ожидания.")

    while True:
        request_data = {}
        try:
            with open(REQUEST_FILE, 'r', encoding='utf-8') as f:
                request_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            time.sleep(5)
            continue

        if request_data.get('status') != 'pending':
            time.sleep(5)
            continue

        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Обнаружен новый запрос на отчет...")

        request_data['status'] = 'processing'
        save_data(REQUEST_FILE, request_data)

        success = False
        error_message = ""
        try:
            access_token = get_access_token()
            if not access_token: raise Exception("Не удалось получить токен доступа Dodo IS.")

            pizzerias = load_pizzerias()
            if not pizzerias: raise Exception(f"Файл '{PIZZERIAS_FILE}' не найден.")
            unit_ids = list(pizzerias.keys())

            end_date_cache = datetime.now()
            start_date_cache = end_date_cache - timedelta(days=30)
            rewards_for_cache = fetch_rewards_data(access_token, unit_ids,
                                                   start_date_cache.strftime('%Y-%m-%dT%H:%M:%S'),
                                                   end_date_cache.strftime('%Y-%m-%dT%H:%M:%S'), is_for_cache=True)
            if rewards_for_cache:
                update_and_save_staff_role_cache(rewards_for_cache)
            else:
                print("  [Предупреждение] Не удалось обновить кэш должностей.")

            start_date_iso = request_data.get('start_date_iso')
            end_date_iso = request_data.get('end_date_iso')

            sales_data_result = fetch_sales_data(access_token, unit_ids, start_date_iso, end_date_iso)
            rewards_data_result = fetch_rewards_data(access_token, unit_ids, start_date_iso, end_date_iso,
                                                     is_for_cache=False)

            if sales_data_result and rewards_data_result:
                generate_and_save_report(pizzerias, sales_data_result, rewards_data_result)
                success = True
            else:
                raise Exception("Один из этапов сбора данных завершился с ошибкой.")

        except Exception as e:
            success = False
            error_message = str(e)
            print(f"[Ошибка] Сбой при генерации отчета: {e}")

        if success:
            request_data['status'] = 'completed'
            request_data['completed_at'] = datetime.now().isoformat()
        else:
            request_data['status'] = 'error'
            request_data['error_message'] = error_message

        save_data(REQUEST_FILE, request_data)
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Обработка завершена: {request_data['status']}")


if __name__ == "__main__":
    main()

